# %%
import camelot
# from camelot import read_pdf
import pyperclip
import os
import re
import sys
import json
import datetime
import pandas as pd
import numpy as np
import requests
import time
import glob
from tqdm import tqdm
import openpyxl
from openpyxl import load_workbook
import camelot
print(camelot.__version__)


# %%

import ctypes
from ctypes.util import find_library
find_library("".join(("gsdll", str(ctypes.sizeof(ctypes.c_voidp) * 8), ".dll")))

# %% Ed Diversity side project
file = r"C:\Users\clutz\Downloads\2024 Rankings and Estimates Report.pdf"

try:
    # Read tables from the PDF file
    tables = camelot.read_pdf(file, pages='all')  # Change 'all' to a specific page number if needed
    print(f"Total tables found: {len(tables)}")

    # Access and display the first table
    if tables:
        first_table = tables[0]
        print(first_table.df)  # Print the DataFrame of the first table
    else:
        print("No tables found in the PDF.")

except Exception as e:
    print(f"An error occurred: {e}")
# %%
tables = camelot.read_pdf(file, pages="43")

for table in tables:
    print(table)

# %% chatgpt code to pull info from html of oklahom committee websites


from bs4 import BeautifulSoup

# Sample HTML (from your provided HTML snippet)
html_content = """
<div class="media-grid media-grid-4-column">

		

			<div class="media-container media-container-portrait last-item-row-1  republican ">

				<a target="_self" href="/members/bill-roemer">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Bill Roemer</div>

								
									<div class="media-overlay-caption-text-line-2">District 31 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2893.jpg)" class="media-thumbnail-image"></div>

					</div>

					<div class="media-caption">
						<div><div class="committee-member-position">Chair</div></div>
					</div>

				</a>

			</div>

			<div class="media-grid-clear-row-1 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2  republican ">

				<a target="_self" href="/members/jack-k-daniels">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Jack K. Daniels</div>

								
									<div class="media-overlay-caption-text-line-2">District 32 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/3013.jpg)" class="media-thumbnail-image"></div>

					</div>

					<div class="media-caption">
						<div><div class="committee-member-position">Vice Chair</div></div>
					</div>

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-3  democrat ">

				<a target="_self" href="/members/daniel-p-troy">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Daniel P. Troy</div>

								
									<div class="media-overlay-caption-text-line-2">District 23 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2948.jpg)" class="media-thumbnail-image"></div>

					</div>

					<div class="media-caption">
						<div><div class="committee-member-position">Ranking Member</div></div>
					</div>

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-3 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-4  democrat ">

				<a target="_self" href="/members/sean-p-brennan">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Sean P. Brennan</div>

								
									<div class="media-overlay-caption-text-line-2">District 14 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2979.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-4 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-5  democrat ">

				<a target="_self" href="/members/richard-dellaquila">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Richard Dell'Aquila</div>

								
									<div class="media-overlay-caption-text-line-2">District 15 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2980.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-5 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-3 last-item-row-6  republican ">

				<a target="_self" href="/members/steve-demetriou">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Steve Demetriou</div>

								
									<div class="media-overlay-caption-text-line-2">District 35 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2986.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-3 media-grid-clear-row-6 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1  republican ">

				<a target="_self" href="/members/thomas-hall">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Thomas Hall</div>

								
									<div class="media-overlay-caption-text-line-2">District 46 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2947.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-4  democrat ">

				<a target="_self" href="/members/dani-isaacsohn">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Dani Isaacsohn</div>

								
									<div class="media-overlay-caption-text-line-2">District 24 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2983.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-4 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-3  republican ">

				<a target="_self" href="/members/angela-n-king">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Angela N. King</div>

								
									<div class="media-overlay-caption-text-line-2">District 84 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2999.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-3 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-5  republican ">

				<a target="_self" href="/members/beth-lear">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Beth Lear</div>

								
									<div class="media-overlay-caption-text-line-2">District 61 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2992.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-5 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1  republican ">

				<a target="_self" href="/members/brian-lorenz">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Brian Lorenz</div>

								
									<div class="media-overlay-caption-text-line-2">District 60 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/3009.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-3 last-item-row-4 last-item-row-6  republican ">

				<a target="_self" href="/members/adam-mathews">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Adam Mathews</div>

								
									<div class="media-overlay-caption-text-line-2">District 56 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2990.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-3 media-grid-clear-row-4 media-grid-clear-row-6 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1  republican ">

				<a target="_self" href="/members/riordan-t-mcclain">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Riordan T. McClain</div>

								
									<div class="media-overlay-caption-text-line-2">District 87 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2876.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2  democrat ">

				<a target="_self" href="/members/elgin-rogers-jr">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Elgin Rogers, Jr.</div>

								
									<div class="media-overlay-caption-text-line-2">District 44 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2973.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-3 last-item-row-5  republican ">

				<a target="_self" href="/members/reggie-stoltzfus">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Reggie Stoltzfus</div>

								
									<div class="media-overlay-caption-text-line-2">District 50 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2898.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-3 media-grid-clear-row-5 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1 last-item-row-2 last-item-row-4  democrat ">

				<a target="_self" href="/members/terrence-upchurch">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Terrence Upchurch</div>

								
									<div class="media-overlay-caption-text-line-2">District 20 <span class="vertical-bar">&nbsp;</span> D</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2882.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 media-grid-clear-row-2 media-grid-clear-row-4 "></div>

		

			<div class="media-container media-container-portrait last-item-row-1  republican ">

				<a target="_self" href="/members/scott-wiggam">

					<div class="media-thumbnail">

						<div class="media-selected-indicator">&nbsp;</div>

						

						<div class="media-overlay-caption">

							

							
								<div class="media-overlay-caption-text-line-1">Scott Wiggam</div>

								
									<div class="media-overlay-caption-text-line-2">District 77 <span class="vertical-bar">&nbsp;</span> R</div>
								

							

						</div>

						<div style="background-image:url(/assets/people/headshots/medium/2842.jpg)" class="media-thumbnail-image"></div>

					</div>

					

				</a>

			</div>

			<div class="media-grid-clear-row-1 "></div>

		

	</div>
"""

# Parse the HTML content
soup = BeautifulSoup(html_content, 'html.parser')

# Find all names using the class selector
names = [name.get_text(strip=True) for name in soup.find_all("div", class_="media-overlay-caption-text-line-1")]

# Display extracted names
print(*names, sep=", ")


# %%

text = """
Representative	G. Andrés Romero	10	D	Chair
Representative	Joy Garratt	29	D	Vice Chair
Representative	Jack Chatfield	67	R	Ranking Member
Representative	Brian G. Baca	8	R	Member
Representative	Candy Spence Ezzell	58	R	Member
Representative	Yanira Gurrola	16	D	Member
Representative	William A. Hall II	3	R	Member
Representative	Susan K. Herrera	41	D	Member
Representative	Raymundo Lara	34	D	Member
Representative	Willie D. Madrid	53	D	Member
Representative	Tanya Mirabal Moya	7	R	Member
Representative	Cristina Parajón	25	D	Member
Representative	Patricia Roybal Caballero	13	D	Member
"""

text_x = text.split('\n')

text_x = [x.replace("\t", " ") for x in text_x]

text_x = [re.sub(r'\s[RD]\s', '', x) for x in text_x]
text_x = [re.sub(r'\d{1,2}', '', x) for x in text_x]
text_x = [x.replace(r'Representative', '') for x in text_x if len(x) != 0]
print(text_x)
all_text = "|".join(text_x)
print(all_text)
parts = [x.rsplit(" ", 1) for x in text_x]
print(parts)


names = []
position = []

for x in text_x:
    split = x.rsplit(" ", 1)
    print(split)
    names.append(split[0])
    position.append(split[-1])
    
comm = pd.DataFrame({"name": names, "position": position})
print(comm)

os.chdir(r'C:\Users\clutz\OneDrive - THE HUNT INSTITUTE\Desktop')
comm.to_csv("nm_com_data.csv")
# %% files for committee updating

os.chdir(r'C:\Users\clutz\OneDrive - THE HUNT INSTITUTE\Desktop\drop data')
files = glob.glob('*')
df = load_workbook(filename = files[0])



sheet_names = df.sheetnames
print(sheet_names)


for file in files:
    print(file.split('_')[0])
    wb = load_workbook(filename = file)
    sheet_names = wb.sheetnames

    df = pd.read_excel(file, sheet_name = sheet_names[0])
    print('################')
    print('House Length')
    print('----------------')
    print(len(df))
    df = pd.read_excel(file, sheet_name = sheet_names[1])
    print('Senate Length')
    print('----------------')
    print(len(df))
    print('\n')
    
    
# %%
text = pyperclip.paste()

print(text)

text_list = text.split(";")

new_list = [str(x).split('a(', 1)[-1] for x in text_list]
syntax_str = ";"
# newer_list = [(str(x)+syntax_str) for x in new_list]


newer_list = list(map(lambda x:"{}{}".format(x, syntax_str), new_list))
[print(x) for x in newer_list]
# %%


names = pyperclip.paste()
print(names)
# %%
names_list = names.replace('\r', '').split('\n')



honors = pyperclip.paste()
honors_list = honors.replace('\r', '').split('\n')
honors_list.append('Rep.')
honors_list = [x for x in honors_list if len(x) > 0]
# %%



honorific = [""]*len(names_list)
print(honorific)

names_df = pd.DataFrame({'name': names_list, 'honorific': honorific})

# %%


for i,name in enumerate(names_df['name']):

    split_name = name.split(' ')

    if len(split_name) == 2:
        continue
    elif len(split_name) >= 3:
        
        #look for honorific
        has_honorific = False
        for j in honors_list:
            # print(j)

            if str(j) in split_name[0]:
                
                honorific = re.findall(f'^{j}', str(name))[0]
                names_df.loc[i,'honorific'] = str(honorific).strip()
                names_df.loc[i,'name'] = name.replace(f'{honorific}', "").strip()

            
        # if has_honorific == True:
        #     str_extract(txt, "[0-9]+")
        #     names_df.loc[i,'honorific'] = str(split_name[0])



# %%


def separate_name(full_name):
     # Normalize the input by stripping whitespace
    full_name = full_name.strip()
    
    # Split the full name by spaces
    parts = full_name.split()
    
    # Initialize variables for first name, middle name, and last name
    first_name = ''
    middle_name = ''
    last_name = ''
    
    # Handle the case with only first and last name
    if len(parts) == 2:
        first_name = parts[0]
        last_name = parts[1]
    else:
        first_name = parts[0]
        last_name = parts[-1]
        
        # Check if the last name is "Jr."
        if last_name.lower() == "jr.":
            last_name = ' '.join(parts[-2:])  # Combine last name and suffix
        else:
            middle_name = ' '.join(parts[1:-1])  # Join middle names/initials if any
    
    # If there's a middle name, include the first letter in the first name
    if middle_name:
        first_name += ' ' + middle_name[0] + '.' if len(middle_name) == 1 else ' ' + middle_name

    return first_name.strip(), last_name.strip()
# %%
names_df['first_name'] = ""
names_df['last_name'] = ""



for i,j in enumerate(names_df['name']):
    print(str(j))
    try:
    
        first_name, last_name = separate_name(str(j))
        names_df.loc[i,'first_name'] = first_name
        names_df.loc[i,'last_name'] = last_name

    except:
         print(str(j))

os.chdir(r'C:\Users\clutz\OneDrive - THE HUNT INSTITUTE\Documents\Data\memberships tables')
names_df.to_csv('names_df.csv')

# %%

# Example usage
name = "John A Doe"
first, last = separate_name(name)
print("First Name:", first)
print("Last Name:", last)
# %%



            print('honorific is ' + str(split_name[0]))
        else:
            result = any(re.search(r'[A-Za-z]{2,}\.', str(x)) for x in split_name)
            if result == True:
                for part in split_name:
                    if re.search(r'[A-Za-z]\.', str(part)):
                        print('middle name is' + str(part))


        
        




# for 






# %%
