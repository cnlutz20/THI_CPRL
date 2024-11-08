# %%
import os, sys, json, datetime, re  # Provides OS-dependent functionality, system-specific parameters, JSON handling, and date/time manipulation
import pandas as pd             # Provides data structures and data analysis tools
import numpy as np              # Supports large, multi-dimensional arrays and matrices
import requests
import time
from tqdm import tqdm
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


from bs4 import BeautifulSoup

# %%

file = r"C:\Users\clutz\OneDrive - THE HUNT INSTITUTE\Documents\Data\legislator data\committees_data.xlsx"

coms = pd.read_excel(file)

in_coms = coms[coms['state'] == "IN"]

# %%
in_coms = in_coms.drop_duplicates(subset=['committee', 'url'], keep='first')
in_coms.reset_index(inplace=True, drop=True)


# %%

# %%

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



# Path to your WebDriver executable (adjust if necessary)
webdriver_path = r"C:\Users\clutz\hunt_env\chrome driver\chromedriver-win64\chromedriver.exe"

# Set up Chrome options
chrome_options = Options()
# chrome_options.add_argument("--headless")  # Run in headless mode (no GUI)

# Set up WebDriver service
service = Service(webdriver_path)

# Initialize WebDriver
driver = webdriver.Chrome(service=service, options=chrome_options)
"C:\Users\clutz\OneDrive - THE HUNT INSTITUTE\Documents\Data\legislator data\IN\IN_committee_data_from_quorum.xlsx"

# %%
# url = in_coms['url'].iloc[0]
# print(in_coms['committee'].iloc[0])

member_dict = {}
for i,url in enumerate(in_coms["url"]):
    driver.get(url)
    time.sleep(5)
    chamber = in_coms.loc[i,'branch']
    # print(response.text)  # Check the raw HTML
    html_from_page = driver.page_source
    soup = BeautifulSoup(html_from_page, 'html.parser')

    member_cards = soup.find_all(class_='MemberCard_memberCardContent__5CHYi')

    if len(member_cards) == 0:
        print(in_coms['committee'].iloc[i])
        break



    members = []
    role = []
    districts = []
    members_df = []
    chambers = []
    for member in member_cards:
        
        name = member.select_one('.MemberCard_memberCardName__AA367 span').text.strip()
        district = member.select_one('.MemberCard_memberCardDistrict__1IKVn').text.strip()
        position = member.select_one('.MemberCard_memberCardPosition__3b90Z p').text.strip()
        
        

        members.extend([name])
        
        districts.extend([district])
        role.extend([position])
        
        # print("chambers print out" + str(chambers))


        try:
            df = pd.DataFrame({"member": members, "district": districts, "position": role})
        except:
            
            print('MISTAKE')
            print("members list length: " + str(len(members)))
            print(members)
            print("roles list length: " + str(len(role)))
            print(role)
            print("district list length: " + str(len(districts)))
            print(districts)

            break

        members_df.append(df)
    in_com_dfs = pd.concat(members_df)
    in_com_dfs.loc[:,'Chamber'] = chamber
    print(in_com_dfs)
    member_dict[in_coms['committee'].iloc[i]] = in_com_dfs



# %%
# Define a function to modify each string
def modify_string(s):
    sen = False
    rep = False
    if re.search(r'Rep\.', s):
        rep = True
    elif re.search(r'Sen\.', s):
        sen = True
    
    if sen == True:
        # Split the string at commas
        parts = re.split('Sen.',s, maxsplit=1)
    elif rep == True:
        parts = re.split('Rep.',s, maxsplit=1)

    elif sen == False and sen == False:
        print('no rep/sen delimeter')
        return None
    
    result = parts[-1]
    # print(type(result))

    result = re.sub(r'\([rdRD]\)',"", str(result))
    result = result.lstrip().rstrip()
    return result

# %%
# Define a function to standardize committee role
def modify_membership(s):

    if re.search(r'Chair', str(s)):
        if 'Vice' in str(s):
            membership = 'Vice Chair' 
        elif re.search(r'^Chair of', str(s)):
            membership = 'Member'
        else:
            membership = 'Chair'

    elif re.search(r'Member',  str(s)):
        membership = 'Member'
    
    return membership

# %%
# Define a function to standardize committee role
def modify_district(s):

    result = re.split(r'District ', str(s))
    result = str(result[-1])
    return result


# %%
committee_results = []
for k,v in member_dict.items():
    df = v
    df = df.drop_duplicates()

    df.loc[:,'member'] = df['member'].apply(modify_string)
    df.loc[:,'position'] = df['position'].apply(modify_membership)
    df.loc[:,'district'] = df['district'].apply(modify_district)
    n = len(df.loc[:, 'member'])
    com_values = [str(k)]*n
    df.loc[:,'committee'] = com_values
    committee_results.append(df)




# %%

committee_info = pd.concat(committee_results, ignore_index=True)


# %% pickling committe info
# %% storing
import pickle


# file_name = os.path.join(os.getcwd(), 'committee_info_df')
# committee_info.to_pickle(file_name)

# dbfile = open('committee_info_df', 'ab')
    
#     # source, destination
# pickle.dump(committee_info, dbfile)                    
# dbfile.close()

committee_info.to_csv('in_committee_info.csv', index=False)

# %% Loading

with open('committee_info_df', 'rb') as dbfile:
    # Step 2: Load the data from the file
    data = pickle.load(dbfile)

in_committee_info = pd.DataFrame(data)

# %%

from openpyxl import load_workbook

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

file = r"C:\Users\clutz\THE HUNT INSTITUTE\The Hunt Institute Team Site - Documents\Development (formerly Grants Management)\!Administrative\Christian\Legislators Data\leg_data_update_10_2024\IN_legislators.xlsx"
names = get_sheetnames_xlsx(file)

# %%


#%% SENATE

in_legs_senate = pd.read_excel(file, sheet_name=names[1])
in_committee_info_senate = in_committee_info[in_committee_info.Chamber.str.contains('senate')]


in_legs_senate_melted = in_legs_senate.melt(id_vars=['State Abbreviation', 'Chamber', 'full title', 'First Name',
       'Last Name', 'Party', 'district', 'tenure', 'leader'], value_vars=[ 'Appropriations overall', 'Education & Career Development',
       'Family & Children Services', 'Health & Provider Services'],
       var_name='committee',  # Name the new column
       value_name='membership_status'  # Optionally, name this column
       )

# Ensure to include 'district', 'variable', and 'position' in the merge
in_committee_info_senate = in_committee_info_senate[['district', 'committee', 'position']]

in_legs_senate_melted['district'] = in_legs_senate_melted['district'].astype(str)

in_legs_senate_melted = in_legs_senate_melted.merge(
    in_committee_info_senate,
    how='left',
    left_on=['district', 'committee'],  # Using the melted 'committee'
    right_on=['district', 'committee']  # Matching it with the 'variable' from in_committee_info_senate
)


in_legs_senate = in_legs_senate_melted.pivot(columns='committee', 
                         values='position',
                         index=['State Abbreviation', 'Chamber', 'full title', 'First Name','Last Name', 'Party', 'district', 'tenure', 'leader']     
       )
in_legs_senate.reset_index(inplace=True)

#%% HOUSE

in_legs_house = pd.read_excel(file, sheet_name=names[0])
in_committee_info_house = in_committee_info[in_committee_info.Chamber.str.contains('house')]


in_legs_house_melted = in_legs_house.melt(id_vars=['State Abbreviation', 'Chamber', 'full title', 'First Name',
       'Last Name', 'Party', 'district', 'tenure', 'leader'], value_vars=['Family, Children & Human Affairs', 'Education', 'Public Health',
       'Appropriations (Ways & Means)'],
       var_name='committee',  # Name the new column
       value_name='membership_status'  # Optionally, name this column
       )

# Ensure to include 'district', 'variable', and 'position' in the merge
in_committee_info_house = in_committee_info_house[['district', 'committee', 'position']]

in_legs_house_melted['district'] = in_legs_house_melted['district'].astype(str)

in_legs_house_melted = in_legs_house_melted.merge(
    in_committee_info_house,
    how='left',
    left_on=['district', 'committee'],  # Using the melted 'committee'
    right_on=['district', 'committee']  # Matching it with the 'variable' from in_committee_info_house
)


in_legs_house = in_legs_house_melted.pivot(columns='committee', 
                         values='position',
                         index=['State Abbreviation', 'Chamber', 'full title', 'First Name','Last Name', 'Party', 'district', 'tenure', 'leader']     
       )
in_legs_house.reset_index(inplace=True)
# %%

export_list = [in_legs_senate, in_legs_house]
in_comm_export_df = pd.concat(export_list)

in_comm_export_df.to_csv('in_committee_export.csv', index=False)

# %%
import openpyxl

filepath = r'C:\Users\clutz\THE HUNT INSTITUTE\The Hunt Institute Team Site - Documents\Development (formerly Grants Management)\!Administrative\Christian\Legislators Data\leg_data_update_10_2024\in_legislator_com_info.xlsx'
wb = openpyxl.Workbook()

wb.save(filepath)

with pd.ExcelWriter("in_legislator_com_info.xlsx", mode="a", engine="openpyxl", if_sheet_exists='replace') as writer:
     
    # use to_excel function and specify the sheet_name and index to 
    # store the dataframe in specified sheet
    in_legs_senate.to_excel(writer, sheet_name="IN_Senate")
    in_legs_house.to_excel(writer, sheet_name="IN_House")


# %%
test.columns
for k in in_committee_info_senate['district']:
    print(type(k))

in_committee_info_senate.columns



result = test.merge(in_committee_info_senate['position'], 
                     how='left', 
                     left_on=['district','committee'], 
                     right_on=['district','variable'])

# %%
for k,v in member_dict.items():
    print('########################')
    print(k)
    print(len(v.at[0,'district']))
    print(v.at[0,'district']).split(' ')
    print('\n')

